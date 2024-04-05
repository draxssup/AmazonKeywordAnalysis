import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
filename = input('Enter name of your file: ')
df = pd.read_csv(filename+'.csv')
title = input('Enter Title: ')
bullet = input('Enter Bullet Points: ')
srch_term = input('Enter Search Term: ')
df['Highlighted'] = 'NULL'
score = 0
for i, x in enumerate(df[df.keys()[0]]):
    if x in title:
        df.at[i, 'Highlighted'] = 'Title'
        score += df.at[i, 'Search']
    if x in bullet and x not in title:
        df.at[i, 'Highlighted'] = 'Bullet'
        score += df.at[i, 'Search']
    if x in srch_term and x not in title and x not in bullet:
        df.at[i, 'Highlighted'] = 'ST'
        score += df.at[i, 'Search']
df.at[len(df),'Search'] = score
filenm = 'datafileoutput.xlsx'
df.to_excel(filenm, index=False)
wb = load_workbook(filenm)
ws = wb.active
colors = {'Title': '00FF00', 'Bullet': 'FFFF00', 'ST': '0000FF'}
for row in ws.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
    for cell in row:
        highlight_type = df.at[cell.row - 2, 'Highlighted']
        if highlight_type in colors:
            cell.fill = PatternFill(start_color=colors[highlight_type], end_color=colors[highlight_type], fill_type="solid")


col_idx = ws.max_column

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
    for cell in row:
        ws.delete_cols(cell.col_idx)

wb.save(filenm)
print(f'file saved as {filenm}')
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import tkinter as tk
from tkinter import messagebox
import os


def submit():
    filename = title.get()
    full_path = os.path.abspath(filename) + '.csv'
    if os.path.exists(full_path):
        print(f'Full Path: {full_path}')
        show_entry_boxes(full_path)
    else:
        messagebox.showerror('Error', f'File not found: {full_path}')


def show_entry_boxes(filename):
    entry_window = tk.Toplevel(root)
    entry_window.title('Enter Details')

    tk.Label(entry_window, text='Title:').grid(row=0, column=0, sticky='w')
    title_entry = tk.Entry(entry_window, width=30)
    title_entry.grid(row=0, column=1, padx=10, pady=5)

    tk.Label(entry_window, text='Bullet Points:').grid(row=1, column=0, sticky='w')
    bullet_entry = tk.Entry(entry_window, width=30)
    bullet_entry.grid(row=1, column=1, padx=10, pady=5)

    tk.Label(entry_window, text='Search Terms:').grid(row=2, column=0, sticky='w')
    search_entry = tk.Entry(entry_window, width=30)
    search_entry.grid(row=2, column=1, padx=10, pady=5)

    tk.Label(entry_window, text='Description:').grid(row=3, column=0, sticky='w')
    description_entry = tk.Entry(entry_window, width=30)
    description_entry.grid(row=3, column=1, padx=10, pady=5)

    def save_details():
        title = title_entry.get().upper()
        bullet = bullet_entry.get().upper()
        search_term = search_entry.get().upper()
        description = description_entry.get().upper()

        df = pd.read_csv(filename)
        df['Highlighted'] = 'NULL'
        score = 0
        for i in range(len(df)):
            x = df.at[i, 'Keyword']
            if x.upper() in title:
                df.at[i, 'Highlighted'] = 'Title'
                score += df.at[i, 'Search']
            if x.upper() in bullet and x.upper() not in title:
                df.at[i, 'Highlighted'] = 'Bullet'
                score += df.at[i, 'Search']
            if x.upper() in search_term and x.upper() not in title and x.upper() not in bullet:
                df.at[i, 'Highlighted'] = 'ST'
                print(description)
                score += df.at[i, 'Search']
            if x.upper() in description and x.upper() not in search_term and x.upper() not in title and x.upper() not in bullet:
                print('ARE CHALRA H BC')
                df.at[i, 'Highlighted'] = 'DSC'
                score += df.at[i, 'Search']
        df.at[len(df), 'Search'] = score

        filenm = 'datafileoutput.xlsx'
        df.to_excel(filenm, index=False)
        wb = load_workbook(filenm)
        ws = wb.active
        colors = {'Title': '00FF00', 'Bullet': 'FFFF00', 'ST': '0000FF', 'DSC':'FF00FF'}
        for row in ws.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                highlight_type = df.at[cell.row - 2, 'Highlighted']
                if highlight_type in colors:
                    cell.fill = PatternFill(start_color=colors[highlight_type], end_color=colors[highlight_type],
                                            fill_type="solid")
        col_idx = ws.max_column
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                ws.delete_cols(cell.col_idx)
        wb.save(filenm)
        print(f'file saved as {filenm}')
        entry_window.destroy()

    save_button = tk.Button(entry_window, text='Save', command=save_details)
    save_button.grid(row=4, column=0, columnspan=2, pady=10)


root = tk.Tk()
root.geometry('500x250')
root.rowconfigure(0, weight=1)
root.rowconfigure(1, weight=1)
root.rowconfigure(2, weight=1)
root.columnconfigure(1, weight=1)

tk.Label(root, text='Enter the name of the file without extension', font=('Arial', 16)).grid(row=0, column=0)
title = tk.Entry(width=45)
title.grid(row=1, column=0, sticky='nw')

submitbtn = tk.Button(root, text='Submit', command=submit)
submitbtn.grid(row=2, column=0, sticky='nw')

root.mainloop()
